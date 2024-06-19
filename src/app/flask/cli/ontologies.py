# Copyright (c) 2021-2024 - Abilian SAS & TCA
#
# SPDX-License-Identifier: AGPL-3.0-only

from __future__ import annotations

from flask.cli import with_appcontext
from flask_super.cli import group
from slugify import slugify

from app.flask.extensions import db
from app.services.taxonomies import create_entry

# Secteurs → `sectors`
# Rubriques → `sections`
# Genres → `genres`
# Type d'info → `topics`
ONTOLOGIES = [
    ("sectors", "NEWS-Secteurs", "level2", 3),
    ("sections", "NEWS-Rubriques", "level1", 4),
    ("topics", "NEWS-Types_d'info", "level2", 4),
    ("genres", "NEWS-Genres", "level1", 4),
    # Not used yet
    ("genres-com", "NEWS-COM-Genres", "level1", 5),
    ("media_type", "Types_de_presse_et_médias", "level1", 3),
    ("organisation_type", "Types_d'organisation", "level1", 4),
    ("job", "Fonctions_du_journalisme", "level1", 2),
]


@group(short_help="Manage ontologies/taxonomies/vocabularies")
def ontologies() -> None:
    pass


@ontologies.command(name="import", short_help="Import ontologies")
@with_appcontext
def import_ontologies() -> None:
    _import_ontologies()


def _import_ontologies() -> None:
    from openpyxl.reader.excel import load_workbook
    from openpyxl.workbook import Workbook

    wb: Workbook = load_workbook("data/Ontologies-12-V2.xlsx")

    for ontology, sheet_name, level, start_at in ONTOLOGIES:
        ws = wb[sheet_name]
        if level == "level1":
            result = process_1_level(ws, start_at)
        else:
            result = process_2_levels(ws, start_at)

        for row in result:
            label = " / ".join(row)
            create_entry(ontology, label)

    db.session.commit()


@ontologies.command(short_help="Initialize index")
@with_appcontext
def dump() -> None:
    pass
    # dump_voc("rubriques", voc.PRESS_SECTIONS, 1)
    # dump_voc("genres", voc.GENRES, 1)
    #
    # dump_voc("competences", voc.COMPETENCIES, 2)
    # dump_voc("secteurs", voc.SECTORS, 2)
    # dump_voc("jobs", voc.JOBS, 2)
    # dump_voc("sujets", voc.MEDIA_TOPICS, 2)


def process_1_level(ws, start_at=3):
    from openpyxl.cell import Cell

    result = []
    for row in ws.iter_rows(min_row=start_at, max_col=3):
        cell: Cell = row[0]
        if not cell.value:
            continue

        result += [(cell.value,)]

    return result


def process_2_levels(ws, start_at):
    from openpyxl.cell import Cell

    result = []
    level1 = ""
    level2 = ""
    for row in ws.iter_rows(min_row=start_at, max_col=3):
        cell: Cell = row[0]
        if not cell.value:
            level1 = ""
            continue

        if not level1:
            level1 = cell.value
        else:
            level2 = cell.value
        result += [(level1, level2)]

    return result


def dump_voc(name, voc, levels=1):
    from openpyxl.workbook import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws

    if levels == 1:
        ws.append(["id", "label"])
    else:
        headers = ["id"]
        for i in range(levels):
            headers.append(f"label {i + 1}")
        ws.append(headers)

    for label in voc:
        id = slugify(label)
        labels = label.split(" / ")
        ws.append([id, *labels])

    normalize(ws)
    wb.save(f"export/{name}.xlsx")


def normalize(sheet):
    def as_text(value):
        if value is None:
            return ""
        return str(value)

    sizes = []
    for line in sheet:
        if not sizes:
            sizes = [len(as_text(cell.value)) for cell in line]
        else:
            for i, cell in enumerate(line):
                sizes[i] = max(sizes[i], len(as_text(cell.value)))
    for i, size in enumerate(sizes):
        col_name = chr(ord("A") + i)
        sheet.column_dimensions[col_name].width = size
