# Copyright (c) 2021-2024, Abilian SAS & TCA
#
# SPDX-License-Identifier: AGPL-3.0-onlyfrom __future__ import annotations

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl.reader.excel import load_workbook

from app.constants import PROFILE_CODES
from app.enums import CommunityEnum, ContactTypeEnum

from .survey_dataclass import (
    Group,
    ModelLoader,
    SurveyCommunities,
    SurveyField,
    SurveyProfile,
)

COLOR_CODES = {
    "FF32CD32": "M",  # Mandatory
    "FFFFA500": "O",  # Optional
    "FFC9211E": "N",  # No
}

ROW_FIELDS = 7
ROW_COMMUNITY = 1
ROW_PROFILE = 2
ROW_PROFILE_CODE = 3
ROW_CONTACT = 4

COL_LABEL = 0
COL_PUBLIC_MAXI = 1
COL_PUBLIC_DEFAULT = 2
COL_PUBLIC_MINI = 3
COL_VALIDATE_CHANGES = 4
COL_MESSAGE = 5
COL_ORGANISATION = 6
COL_ID = 7
COL_TYPE = 8
COL_COMMENT = 9
COL_PROFILE = 10


KYC_COMMUNITY_TO_ENUM: dict[str, CommunityEnum] = {  # type: ignore
    "press & media": CommunityEnum.PRESS_MEDIA,
    "press relations": CommunityEnum.COMMUNICANTS,
    "leaders & experts": CommunityEnum.LEADERS_EXPERTS,
    "transformers": CommunityEnum.TRANSFORMERS,
    "academics": CommunityEnum.ACADEMICS,
}


def kyc_community_to_enum(kyc_community: str) -> CommunityEnum:
    return KYC_COMMUNITY_TO_ENUM[kyc_community.lower().strip()]


def kyc_contact_type_to_enum(kyc_contact_type: str) -> ContactTypeEnum:
    key = kyc_contact_type.upper().strip()
    return ContactTypeEnum[key]


def _cell_to_bool(cell) -> bool:
    return bool((cell.value or "").strip())


class XLSParser(ModelLoader):
    def __init__(self) -> None:
        self.survey_fields: dict[str, SurveyField] = {}
        self.survey_profiles: list[SurveyProfile] = []
        self.survey_communities: SurveyCommunities = SurveyCommunities()

    def parse(self, source: Path | str) -> None:
        wb = load_workbook(source)
        ws = wb.active

        rows = list(ws.iter_rows())

        self.read_fields(rows)
        self.read_profiles(rows)
        self.parse_fields(rows)
        self.build_communities()

    @property
    def model(self) -> dict[str, Any]:
        return {
            "communities": self.survey_communities,
            "survey_fields": self.survey_fields,
            "profiles": self.survey_profiles,
        }

    def parse_fields(self, rows) -> None:
        field_i = 0
        for row in rows[ROW_FIELDS:]:
            field_description = row[COL_LABEL].value
            field_name = row[COL_ID].value
            field_type = row[COL_TYPE].value
            if field_type == "title":
                # it's a group title
                for i, _cell in enumerate(row[COL_PROFILE:]):
                    if i >= len(self.survey_profiles):
                        break
                    profile = self.survey_profiles[i]
                    profile.groups.append(Group(label=field_description))
                continue
            if not field_name:
                continue
            field_i += 1
            field_id = f"F{field_i:03}"
            for i, cell in enumerate(row[COL_PROFILE:]):
                if i >= len(self.survey_profiles):
                    break
                profile = self.survey_profiles[i]
                code = COLOR_CODES.get(cell.fill.fgColor.rgb, "?")
                if code == "N":
                    continue
                profile_current_group = profile.groups[-1]
                field = self.survey_fields[field_id]
                profile_current_group.survey_fields.append((field, code))

    def dump(self):
        print("# Modèle KYC")
        print()
        self.dump_fields()
        self.dump_profiles()

    def read_fields(self, rows) -> None:
        field_i = 0
        for row in rows[ROW_FIELDS:]:
            field_description = row[COL_LABEL].value
            field_public_maxi = _cell_to_bool(row[COL_PUBLIC_MAXI])
            field_public_default = _cell_to_bool(row[COL_PUBLIC_DEFAULT])
            field_public_mini = _cell_to_bool(row[COL_PUBLIC_MINI])
            field_validate_changes = _cell_to_bool(row[COL_VALIDATE_CHANGES])
            field_message = row[COL_MESSAGE].value
            field_is_organisation = _cell_to_bool(row[COL_ORGANISATION])
            field_name = row[COL_ID].value
            field_type = row[COL_TYPE].value
            if not field_name or field_type == "title":
                continue
            field_i += 1
            field_id = f"F{field_i:03}"
            field = SurveyField(
                id=field_id,
                name=field_name,
                public_maxi=field_public_maxi,
                public_default=field_public_default,
                public_mini=field_public_mini,
                validate_changes=field_validate_changes,
                is_organisation=field_is_organisation,
                type=field_type,
                description=field_description,
                upper_message=field_message,
            )
            self.survey_fields[field_id] = field

    def read_profiles(self, rows) -> None:
        for i, cell in enumerate(rows[ROW_PROFILE][COL_PROFILE:]):
            description = cell.value
            if not description or not description.strip():
                break
            id = f"P{i + 1:03}"
            profile_code = rows[ROW_PROFILE_CODE][cell.column - 1].value
            if profile_code not in PROFILE_CODES:
                raise ValueError(f"Bad profile code {profile_code!r}")
            profile = SurveyProfile(id=id, description=description, code=profile_code)
            self.survey_profiles.append(profile)
        nb_profiles = len(self.survey_profiles)
        community = ""
        for i, cell in enumerate(rows[ROW_COMMUNITY][COL_PROFILE:]):
            if i >= nb_profiles:
                break
            if cell.value:
                community = kyc_community_to_enum(cell.value)
            profile = self.survey_profiles[i]
            profile.community = community

        for i, cell in enumerate(rows[ROW_CONTACT][COL_PROFILE:]):
            if i >= nb_profiles:
                break
            if cell.value:
                contact_type = kyc_contact_type_to_enum(cell.value)
            profile = self.survey_profiles[i]
            profile.contact_type = contact_type

    def build_communities(self) -> None:
        self.survey_communities = SurveyCommunities()
        for profile in self.survey_profiles:
            self.survey_communities.add_profile(profile)

    def dump_fields(self):
        print()
        print("## Fields")
        print()
        for field in self.survey_fields.values():
            print(f"### {field.id}")
            print()
            print(field.description)
            print()

    def dump_profiles(self):
        print()
        print("## Profiles")
        print()
        for profile in self.survey_profiles:
            print(f"### {profile.community} - {profile.id}")
            print()
            print(profile.description)
            print()
            for group in profile.groups:
                print(f"grouo: {group.label}")
                for field, code in group.survey_fields:
                    if code in {"?", "N"}:
                        continue
                    print(f"- {field.id}({code}): {field.description}")
            print()
